from __future__ import annotations

import logging
import re
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.database import SessionLocal
from app.models.application import Application
from app.models.application_roadmap import (
    ApplicationRoadmapDetail,
    ApplicationRoadmapImport,
    ApplicationRoadmapResource,
    ApplicationRoadmapTeam,
    RoadmapEnvironment,
    RoadmapPhase,
    RoadmapResource,
    RoadmapTeam,
)


logger = logging.getLogger(__name__)


# ============================================================
# COLUMN HEADERS
# ============================================================

COLUMN_ALIASES = {
    "activity_number": {
        "#",
        "no",
        "no.",
        "sr no",
        "sr. no.",
        "s.no",
        "serial number",
    },
    "activity": {
        "activity",
        "activities",
        "task",
        "tasks",
    },
    "responsible_team": {
        "resp. team",
        "resp team",
        "responsible team",
    },
    "support_team": {
        "support",
        "support team",
    },
    "assigned_resource": {
        "assigned resource",
        "assigned resources",
        "assigned to",
        "resource",
    },
    "status": {
        "status",
    },
    "planned_start_date": {
        "planned start date",
        "planned start",
    },
    "planned_end_date": {
        "planned end date",
        "planned end",
    },
    "actual_start_date": {
        "actual start date",
        "actual start",
    },
    "actual_end_date": {
        "actual end date",
        "actual end",
    },
    "remarks": {
        "remarks",
        "remark",
        "comments",
        "comment",
    },
}


# ============================================================
# PHASES
# ============================================================

PHASE_MAPPING = {
    "assessment industrialization": {
        "name": "ASSESSMENT",
        "display_name": "Assessment Industrialization",
        "display_order": 1,
    },
    "assessment industrilization": {
        "name": "ASSESSMENT",
        "display_name": "Assessment Industrialization",
        "display_order": 1,
    },
    "assessment": {
        "name": "ASSESSMENT",
        "display_name": "Assessment",
        "display_order": 1,
    },
    "prepare phase": {
        "name": "PREPARE",
        "display_name": "Prepare Phase",
        "display_order": 2,
    },
    "migrate phase": {
        "name": "MIGRATE",
        "display_name": "Migrate Phase",
        "display_order": 3,
    },
    "stability/decomm phase": {
        "name": "STABILITY_DECOMM",
        "display_name": "Stability/Decomm Phase",
        "display_order": 4,
    },
}


# ============================================================
# ENVIRONMENTS
# ============================================================

ENVIRONMENT_MAPPING = {
    "general": {
        "name": "GENERAL",
        "display_name": "General Tasks",
        "display_order": 1,
    },
    "dev": {
        "name": "DEV",
        "display_name": "DEV",
        "display_order": 2,
    },
    "qa": {
        "name": "QA",
        "display_name": "QA",
        "display_order": 3,
    },
    "uat/am": {
        "name": "UAT_AM",
        "display_name": "UAT/AM",
        "display_order": 4,
    },
    "mnt/e1": {
        "name": "MNT_E1",
        "display_name": "MNT/E1",
        "display_order": 5,
    },
    "bench": {
        "name": "BENCH",
        "display_name": "BENCH",
        "display_order": 6,
    },
    "staging": {
        "name": "STAGING",
        "display_name": "STAGING",
        "display_order": 7,
    },
    "int": {
        "name": "INT",
        "display_name": "INT",
        "display_order": 8,
    },
    "pre-prod": {
        "name": "PRE_PROD",
        "display_name": "PRE-PROD",
        "display_order": 9,
    },
    "prod": {
        "name": "PROD",
        "display_name": "PROD",
        "display_order": 10,
    },
}


STATUS_MAPPING = {
    "done": "DONE",
    "to do": "TO_DO",
    "todo": "TO_DO",
    "in progress": "IN_PROGRESS",
    "not required": "NOT_REQUIRED",
    "blocked": "BLOCKED",
    "on hold": "ON_HOLD",
    "cancelled": "CANCELLED",
}


def utc_now() -> datetime:
    return datetime.now(timezone.utc)


def clean_text(value: Any) -> str | None:
    if value is None:
        return None

    text = str(value)
    text = text.replace("\r", " ")
    text = text.replace("\n", " ")
    text = re.sub(r"\s+", " ", text).strip()

    return text or None


def normalize_text(value: Any) -> str:
    text = clean_text(value)

    if not text:
        return ""

    return text.casefold()


def normalize_master_name(value: str) -> str:
    value = normalize_text(value)
    value = re.sub(r"\s+", " ", value)
    value = re.sub(r"\s*/\s*", "/", value)

    return value.strip()


def parse_integer(value: Any) -> int | None:
    if value is None or isinstance(value, bool):
        return None

    if isinstance(value, int):
        return value

    if isinstance(value, float):
        return int(value)

    text = clean_text(value)

    if not text:
        return None

    match = re.search(r"\d+", text)

    if not match:
        return None

    return int(match.group())


def parse_excel_date(value: Any) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    text = clean_text(value)

    if not text:
        return None

    formats = (
        "%d-%b-%y",
        "%d-%b-%Y",
        "%d/%m/%Y",
        "%d/%m/%y",
        "%d-%m-%Y",
        "%d-%m-%y",
        "%Y-%m-%d",
    )

    for date_format in formats:
        try:
            return datetime.strptime(
                text,
                date_format,
            ).date()
        except ValueError:
            continue

    return None


def normalize_status(value: Any) -> str | None:
    status = normalize_text(value)

    if not status:
        return None

    return STATUS_MAPPING.get(
        status,
        status.upper().replace(" ", "_"),
    )


def build_column_mapping(
    row_values: list[Any],
) -> dict[str, int]:
    mapping: dict[str, int] = {}

    for column_index, value in enumerate(row_values):
        header = normalize_text(value)

        if not header:
            continue

        for field_name, aliases in COLUMN_ALIASES.items():
            normalized_aliases = {
                normalize_text(alias)
                for alias in aliases
            }

            if header in normalized_aliases:
                mapping[field_name] = column_index
                break

    return mapping


def find_header_row(
    worksheet: Worksheet,
) -> tuple[int, dict[str, int]]:
    scan_limit = min(worksheet.max_row, 50)

    for row_number in range(1, scan_limit + 1):
        row_values = [
            worksheet.cell(
                row=row_number,
                column=column_number,
            ).value
            for column_number in range(
                1,
                worksheet.max_column + 1,
            )
        ]

        mapping = build_column_mapping(row_values)

        if (
            "activity" in mapping
            and "status" in mapping
        ):
            return row_number, mapping

    raise ValueError(
        f"Roadmap header was not found in "
        f"worksheet '{worksheet.title}'."
    )


def get_mapped_value(
    row_values: list[Any],
    mapping: dict[str, int],
    field_name: str,
) -> Any:
    column_index = mapping.get(field_name)

    if column_index is None:
        return None

    if column_index >= len(row_values):
        return None

    return row_values[column_index]


def detect_phase(
    row_values: list[Any],
) -> dict[str, Any] | None:
    row_text = " ".join(
        clean_text(value) or ""
        for value in row_values
    )

    normalized = normalize_text(row_text)

    for phrase in sorted(
        PHASE_MAPPING.keys(),
        key=len,
        reverse=True,
    ):
        if phrase in normalized:
            return PHASE_MAPPING[phrase]

    return None


def detect_environment(
    row_values: list[Any],
) -> dict[str, Any] | None:
    row_text = " ".join(
        clean_text(value) or ""
        for value in row_values
    )

    match = re.search(
        r"environment\s*\(\s*([^)]+?)\s*\)",
        row_text,
        flags=re.IGNORECASE,
    )

    if match:
        environment_name = normalize_master_name(
            match.group(1)
        )

        return ENVIRONMENT_MAPPING.get(
            environment_name
        )

    normalized = normalize_text(row_text)

    if (
        "general tasks" in normalized
        or "cluster independent tasks" in normalized
    ):
        return ENVIRONMENT_MAPPING["general"]

    return None


def detect_section(
    row_values: list[Any],
) -> str | None:
    row_text = " ".join(
        clean_text(value) or ""
        for value in row_values
    )

    normalized = normalize_text(row_text)

    if (
        "general tasks" in normalized
        or "cluster independent tasks" in normalized
    ):
        return clean_text(row_text)

    return None


def split_multiple_values(
    value: str | None,
) -> list[str]:
    if not value:
        return []

    parts = re.split(
        r"\s*(?:/|;|,|\n)\s*",
        value,
    )

    results: list[str] = []
    seen: set[str] = set()

    for part in parts:
        cleaned = clean_text(part)

        if not cleaned:
            continue

        normalized = normalize_master_name(cleaned)

        if normalized in seen:
            continue

        seen.add(normalized)
        results.append(cleaned)

    return results


def get_or_create_phase(
    db: Session,
    cache: dict[str, RoadmapPhase],
    phase_data: dict[str, Any],
) -> RoadmapPhase:
    name = phase_data["name"]

    if name in cache:
        return cache[name]

    phase = (
        db.query(RoadmapPhase)
        .filter(RoadmapPhase.name == name)
        .one_or_none()
    )

    if phase is None:
        phase = RoadmapPhase(
            name=name,
            display_name=phase_data["display_name"],
            display_order=phase_data["display_order"],
        )

        db.add(phase)
        db.flush()

    cache[name] = phase
    return phase


def get_or_create_environment(
    db: Session,
    cache: dict[str, RoadmapEnvironment],
    environment_data: dict[str, Any],
) -> RoadmapEnvironment:
    name = environment_data["name"]

    if name in cache:
        return cache[name]

    environment = (
        db.query(RoadmapEnvironment)
        .filter(RoadmapEnvironment.name == name)
        .one_or_none()
    )

    if environment is None:
        environment = RoadmapEnvironment(
            name=name,
            display_name=environment_data["display_name"],
            display_order=environment_data["display_order"],
        )

        db.add(environment)
        db.flush()

    cache[name] = environment
    return environment


def get_or_create_team(
    db: Session,
    cache: dict[str, RoadmapTeam],
    team_name: str,
) -> RoadmapTeam:
    normalized_name = normalize_master_name(team_name)

    if normalized_name in cache:
        return cache[normalized_name]

    team = (
        db.query(RoadmapTeam)
        .filter(
            RoadmapTeam.normalized_name
            == normalized_name
        )
        .one_or_none()
    )

    if team is None:
        team = RoadmapTeam(
            name=team_name,
            normalized_name=normalized_name,
        )

        db.add(team)
        db.flush()

    cache[normalized_name] = team
    return team


def get_or_create_resource(
    db: Session,
    cache: dict[str, RoadmapResource],
    resource_name: str,
) -> RoadmapResource:
    normalized_name = normalize_master_name(
        resource_name
    )

    if normalized_name in cache:
        return cache[normalized_name]

    resource = (
        db.query(RoadmapResource)
        .filter(
            RoadmapResource.normalized_name
            == normalized_name
        )
        .one_or_none()
    )

    if resource is None:
        resource = RoadmapResource(
            name=resource_name,
            normalized_name=normalized_name,
        )

        db.add(resource)
        db.flush()

    cache[normalized_name] = resource
    return resource


def add_team_assignments(
    db: Session,
    detail: ApplicationRoadmapDetail,
    raw_value: str | None,
    role: str,
    team_cache: dict[str, RoadmapTeam],
) -> None:
    team_names = split_multiple_values(raw_value)

    for index, team_name in enumerate(team_names):
        team = get_or_create_team(
            db=db,
            cache=team_cache,
            team_name=team_name,
        )

        db.add(
            ApplicationRoadmapTeam(
                roadmap_detail_id=detail.id,
                team_id=team.id,
                role=role,
                is_primary=index == 0,
            )
        )


def add_resource_assignments(
    db: Session,
    detail: ApplicationRoadmapDetail,
    raw_value: str | None,
    resource_cache: dict[str, RoadmapResource],
) -> None:
    resource_names = split_multiple_values(
        raw_value
    )

    for index, resource_name in enumerate(
        resource_names
    ):
        resource = get_or_create_resource(
            db=db,
            cache=resource_cache,
            resource_name=resource_name,
        )

        db.add(
            ApplicationRoadmapResource(
                roadmap_detail_id=detail.id,
                resource_id=resource.id,
                is_primary=index == 0,
            )
        )


def process_roadmap_excel_background(
    import_id: int,
    application_id: int,
    file_path: str,
    replace_existing: bool = False,
) -> None:
    """
    Complete background task.

    This function:
    1. Opens the Excel file.
    2. Detects headers.
    3. Detects phases and environments.
    4. Extracts activity values.
    5. Creates teams and resources.
    6. Saves roadmap details for the application.
    7. Updates import status.
    """

    db = SessionLocal()

    try:
        import_record = (
            db.query(ApplicationRoadmapImport)
            .filter(
                ApplicationRoadmapImport.id == import_id
            )
            .one_or_none()
        )

        if import_record is None:
            raise ValueError(
                f"Import record {import_id} was not found."
            )

        application = (
            db.query(Application)
            .filter(
                Application.id == application_id
            )
            .one_or_none()
        )

        if application is None:
            raise ValueError(
                f"Application {application_id} was not found."
            )

        import_record.import_status = "PROCESSING"
        import_record.started_at = utc_now()
        import_record.error_details = None
        db.commit()

        if replace_existing:
            (
                db.query(ApplicationRoadmapDetail)
                .filter(
                    ApplicationRoadmapDetail.application_id
                    == application_id
                )
                .delete(synchronize_session=False)
            )

            db.commit()

        workbook = load_workbook(
            filename=file_path,
            data_only=True,
            read_only=False,
        )

        phase_cache: dict[str, RoadmapPhase] = {}
        environment_cache: dict[
            str,
            RoadmapEnvironment,
        ] = {}
        team_cache: dict[str, RoadmapTeam] = {}
        resource_cache: dict[
            str,
            RoadmapResource,
        ] = {}

        total_rows = 0
        imported_rows = 0
        skipped_rows = 0
        failed_rows = 0
        global_display_order = 0

        for worksheet in workbook.worksheets:
            try:
                (
                    header_row_number,
                    column_mapping,
                ) = find_header_row(worksheet)

            except ValueError:
                logger.warning(
                    "Skipping sheet '%s': roadmap header not found.",
                    worksheet.title,
                )
                continue

            current_phase: RoadmapPhase | None = None
            current_environment: (
                RoadmapEnvironment | None
            ) = None
            current_section: str | None = None

            for row_number in range(
                header_row_number + 1,
                worksheet.max_row + 1,
            ):
                row_values = [
                    worksheet.cell(
                        row=row_number,
                        column=column_number,
                    ).value
                    for column_number in range(
                        1,
                        worksheet.max_column + 1,
                    )
                ]

                if not any(
                    clean_text(value)
                    for value in row_values
                ):
                    continue

                phase_data = detect_phase(row_values)

                if phase_data:
                    current_phase = get_or_create_phase(
                        db=db,
                        cache=phase_cache,
                        phase_data=phase_data,
                    )

                    current_environment = None
                    current_section = None
                    continue

                environment_data = detect_environment(
                    row_values
                )

                if environment_data:
                    current_environment = (
                        get_or_create_environment(
                            db=db,
                            cache=environment_cache,
                            environment_data=environment_data,
                        )
                    )

                    detected_section = detect_section(
                        row_values
                    )

                    if detected_section:
                        current_section = detected_section

                    continue

                detected_section = detect_section(
                    row_values
                )

                if detected_section:
                    current_section = detected_section

                    if current_environment is None:
                        current_environment = (
                            get_or_create_environment(
                                db=db,
                                cache=environment_cache,
                                environment_data=(
                                    ENVIRONMENT_MAPPING[
                                        "general"
                                    ]
                                ),
                            )
                        )

                    continue

                activity = clean_text(
                    get_mapped_value(
                        row_values,
                        column_mapping,
                        "activity",
                    )
                )

                if not activity:
                    continue

                total_rows += 1

                if current_phase is None:
                    skipped_rows += 1

                    logger.warning(
                        "Skipped sheet '%s', row %s: "
                        "phase not detected.",
                        worksheet.title,
                        row_number,
                    )
                    continue

                if current_environment is None:
                    current_environment = (
                        get_or_create_environment(
                            db=db,
                            cache=environment_cache,
                            environment_data=(
                                ENVIRONMENT_MAPPING["general"]
                            ),
                        )
                    )

                try:
                    with db.begin_nested():
                        global_display_order += 1

                        responsible_team = clean_text(
                            get_mapped_value(
                                row_values,
                                column_mapping,
                                "responsible_team",
                            )
                        )

                        support_team = clean_text(
                            get_mapped_value(
                                row_values,
                                column_mapping,
                                "support_team",
                            )
                        )

                        assigned_resource = clean_text(
                            get_mapped_value(
                                row_values,
                                column_mapping,
                                "assigned_resource",
                            )
                        )

                        detail = ApplicationRoadmapDetail(
                            application_id=application_id,
                            phase_id=current_phase.id,
                            environment_id=(
                                current_environment.id
                            ),
                            import_batch_id=import_id,
                            section_name=current_section,
                            activity_number=parse_integer(
                                get_mapped_value(
                                    row_values,
                                    column_mapping,
                                    "activity_number",
                                )
                            ),
                            activity=activity,
                            status=normalize_status(
                                get_mapped_value(
                                    row_values,
                                    column_mapping,
                                    "status",
                                )
                            ),
                            planned_start_date=(
                                parse_excel_date(
                                    get_mapped_value(
                                        row_values,
                                        column_mapping,
                                        "planned_start_date",
                                    )
                                )
                            ),
                            planned_end_date=(
                                parse_excel_date(
                                    get_mapped_value(
                                        row_values,
                                        column_mapping,
                                        "planned_end_date",
                                    )
                                )
                            ),
                            actual_start_date=(
                                parse_excel_date(
                                    get_mapped_value(
                                        row_values,
                                        column_mapping,
                                        "actual_start_date",
                                    )
                                )
                            ),
                            actual_end_date=(
                                parse_excel_date(
                                    get_mapped_value(
                                        row_values,
                                        column_mapping,
                                        "actual_end_date",
                                    )
                                )
                            ),
                            remarks=clean_text(
                                get_mapped_value(
                                    row_values,
                                    column_mapping,
                                    "remarks",
                                )
                            ),
                            display_order=(
                                global_display_order
                            ),
                            raw_responsible_team=(
                                responsible_team
                            ),
                            raw_support_team=support_team,
                            raw_assigned_resource=(
                                assigned_resource
                            ),
                            source_sheet_name=(
                                worksheet.title
                            ),
                            source_row_number=row_number,
                        )

                        db.add(detail)
                        db.flush()

                        add_team_assignments(
                            db=db,
                            detail=detail,
                            raw_value=responsible_team,
                            role="RESPONSIBLE",
                            team_cache=team_cache,
                        )

                        add_team_assignments(
                            db=db,
                            detail=detail,
                            raw_value=support_team,
                            role="SUPPORT",
                            team_cache=team_cache,
                        )

                        add_resource_assignments(
                            db=db,
                            detail=detail,
                            raw_value=assigned_resource,
                            resource_cache=resource_cache,
                        )

                    imported_rows += 1

                    if imported_rows % 200 == 0:
                        db.commit()

                except Exception as row_error:
                    failed_rows += 1

                    logger.exception(
                        "Failed importing sheet '%s', "
                        "row %s: %s",
                        worksheet.title,
                        row_number,
                        row_error,
                    )

        workbook.close()

        db.commit()

        import_record = (
            db.query(ApplicationRoadmapImport)
            .filter(
                ApplicationRoadmapImport.id == import_id
            )
            .one()
        )

        import_record.total_rows = total_rows
        import_record.imported_rows = imported_rows
        import_record.skipped_rows = skipped_rows
        import_record.failed_rows = failed_rows
        import_record.completed_at = utc_now()

        if failed_rows or skipped_rows:
            import_record.import_status = (
                "PARTIALLY_COMPLETED"
            )
        else:
            import_record.import_status = "COMPLETED"

        db.commit()

    except Exception as error:
        db.rollback()

        logger.exception(
            "Roadmap import %s failed.",
            import_id,
        )

        import_record = (
            db.query(ApplicationRoadmapImport)
            .filter(
                ApplicationRoadmapImport.id == import_id
            )
            .one_or_none()
        )

        if import_record:
            import_record.import_status = "FAILED"
            import_record.error_details = str(error)
            import_record.completed_at = utc_now()
            db.commit()

    finally:
        db.close()

        try:
            Path(file_path).unlink(missing_ok=True)
        except OSError:
            logger.warning(
                "Could not remove temporary file: %s",
                file_path,
            )